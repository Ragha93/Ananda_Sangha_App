"""
Excel file management utilities for Kriyaban Meditation Sangha application
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from pathlib import Path


class ExcelManager:
    """Manages Excel file operations for the application"""

    SHEET_NAMES = {
        'Donations': ['ID', 'Name', 'Email', 'Phone', 'Amount',
                      'Aadhar', 'PAN', 'Address', 'Date', 'Time'],
        'Meditation': ['ID', 'Name', 'Email', 'Phone', 'Referral_Source',
                       'Referral_Details', 'Meditation_Level', 'Course_Fee',
                       'Aadhar', 'PAN', 'Address', 'Enrollment_Date', 'Enrollment_Time'],
        'Boutique': ['ID', 'Item_Name', 'Price', 'Payment_Method', 'Date', 'Time'],
        'Books': ['ID', 'Book_Name', 'Author', 'Price', 'Quantity', 'Customer_Name',
                 'Email', 'Phone', 'Date', 'Time'],
        'Inventory': ['Item_ID', 'Item_Name', 'Category', 'Current_Stock',
                     'Reorder_Level', 'Last_Updated']
    }

    def __init__(self):
        self.file_path = None
        self.workbook = None

    def create_new_file(self, file_path):
        """Create a new Excel file with all required sheets"""
        try:
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)  # Remove default sheet

            # Create sheets with headers
            for sheet_name, headers in self.SHEET_NAMES.items():
                ws = workbook.create_sheet(sheet_name)
                ws.append(headers)

                # Style headers
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            workbook.save(file_path)
            self.file_path = file_path
            self.workbook = workbook
            return True
        except Exception as e:
            print(f"Error creating Excel file: {e}")
            return False

    def load_file(self, file_path):
        """Load an existing Excel file"""
        try:
            if not os.path.exists(file_path):
                return False

            self.workbook = openpyxl.load_workbook(file_path)
            self.file_path = file_path
            return True
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return False

    def save_file(self):
        """Save workbook. Retries up to 5× if file is locked, then saves to a timestamped backup."""
        import time
        if not self.workbook or not self.file_path:
            return False

        for attempt in range(5):
            try:
                self.workbook.save(self.file_path)
                return True
            except PermissionError:
                if attempt < 4:
                    time.sleep(0.5)   # wait 0.5 s and retry
            except Exception as e:
                print(f"Save error: {e}")
                return False

        # Still locked after 2.5 s → save to a timestamped backup in the same folder
        try:
            fp = Path(self.file_path)
            backup = fp.parent / f"{fp.stem}_{datetime.now().strftime('%H%M%S')}.xlsx"
            self.workbook.save(str(backup))
            print(f"Original file locked — data saved to backup: {backup}")
            return True
        except Exception as e:
            print(f"Backup save failed: {e}")
            return False

    def add_donation(self, name, email, phone, amount,
                     aadhar='', address='', pan=''):
        """Add a donation record. Returns (True, receipt_no) where receipt_no is sequential."""
        try:
            unique_id = self.generate_unique_id(email, phone)
            ws = self.workbook['Donations']

            if self.check_duplicate(ws, unique_id):
                return False, "A donation record already exists for this person"

            receipt_no = ws.max_row      # sequential: 1 for first record, 2 for second, etc.
            now = datetime.now()
            ws.append([unique_id, name, email, phone, amount,
                        aadhar, pan, address,
                        now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")])
            self.save_file()
            return True, receipt_no
        except Exception as e:
            return False, str(e)

    def add_meditation(self, name, email, phone, referral_source, referral_details,
                       meditation_level, course_fee,
                       aadhar='', address='', pan=''):
        """Add a meditation enrollment. Returns (True, receipt_no). Blocks duplicate unique_id."""
        try:
            unique_id = self.generate_unique_id(email, phone)
            ws = self.workbook['Meditation']

            if self.check_duplicate(ws, unique_id):
                return False, "A meditation enrollment already exists for this person"

            receipt_no = ws.max_row
            now = datetime.now()
            ws.append([unique_id, name, email, phone, referral_source, referral_details,
                        meditation_level, course_fee,
                        aadhar, pan, address,
                        now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")])
            self.save_file()
            return True, receipt_no
        except Exception as e:
            return False, str(e)

    def add_boutique(self, item_name, price, payment_method):
        """Add a boutique purchase. Returns (True, receipt_no)."""
        try:
            ws = self.workbook['Boutique']
            receipt_no = ws.max_row
            now = datetime.now()
            btq_id = f"BTQ_{now.strftime('%Y%m%d_%H%M%S')}"
            ws.append([btq_id, item_name, price, payment_method,
                        now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")])
            self.save_file()
            return True, receipt_no
        except Exception as e:
            return False, str(e)

    def check_duplicate(self, worksheet, unique_id):
        """Check if a record with the given unique_id already exists"""
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row and row[0] == unique_id:
                return True
        return False

    @staticmethod
    def generate_unique_id(email, phone):
        """Generate unique ID from email and phone (last 4 digits)"""
        phone_last_4 = str(phone)[-4:] if phone else '0000'
        email_part = email if email else 'nomail'
        return f"{email_part}_{phone_last_4}"

    def get_sheet(self, sheet_name):
        """Get a worksheet by name"""
        try:
            return self.workbook[sheet_name] if self.workbook else None
        except:
            return None

    def get_all_records(self, sheet_name):
        """Get all records from a sheet"""
        try:
            ws = self.get_sheet(sheet_name)
            if not ws:
                return []

            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:  # Skip empty rows
                    records.append(row)
            return records
        except Exception as e:
            print(f"Error retrieving records: {e}")
            return []
